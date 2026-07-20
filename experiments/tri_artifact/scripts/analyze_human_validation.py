from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REJECT = "REJECT"
CLARIFY = "CLARIFY"
INVALID = "INVALID_BOUND_ENTITY"

CHINESE_COLUMNS = {
    "题目编号（勿修改）": "item_id",
    "英文任务指令": "instruction",
    "初始状态（每个对象一行）": "initial_state",
    "刷新后状态（每个对象一行）": "refreshed_state",
    "动作前置条件（每项一行）": "action_schema",
    "可选答案（每项一行）": "candidate_ids",
    "你的答案（ID/REJECT/CLARIFY）": "response",
    "信心1到5（可选）": "confidence_1_to_5",
    "备注（可选）": "comment",
}


def normalize_response(value: Any) -> str:
    response = "" if value is None else str(value).strip()
    if response.upper() == CLARIFY:
        return CLARIFY
    if response.upper() in {REJECT, INVALID}:
        return REJECT
    return response


def parse_candidates(value: Any) -> set[str]:
    text = "" if value is None else str(value)
    if "\n" in text:
        values = [line.strip().lstrip("•").strip() for line in text.splitlines()]
    else:
        values = [part.strip() for part in text.split("|")]
    return {value for value in values if value and value not in {REJECT, CLARIFY}}


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def load_xlsx(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(f"Expected one worksheet in {path}, found {workbook.sheetnames}")
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty workbook: {path}")
    headers = [CHINESE_COLUMNS.get(str(value).strip(), str(value).strip()) for value in rows[0]]
    return [
        {headers[index]: "" if value is None else str(value) for index, value in enumerate(row)}
        for row in rows[1:]
        if any(value is not None and str(value).strip() for value in row)
    ]


def load_form(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        return load_xlsx(path)
    return load_csv(path)


def semantic_label(response: str, key: dict[str, str]) -> str:
    if response == CLARIFY:
        return CLARIFY
    if response == REJECT:
        return REJECT
    pre = key["pre_refresh_target"]
    post = key["post_refresh_target"]
    if response == pre == post:
        return "ORIGINAL_AND_REFRESHED"
    if response == pre:
        return "ORIGINAL_ENTITY"
    if response == post:
        return "REFRESHED_SELECTOR"
    return "OTHER_ENTITY"


def wilson(successes: int, total: int, z: float = 1.959963984540054) -> list[float]:
    if total == 0:
        return [float("nan"), float("nan")]
    p = successes / total
    denominator = 1 + z * z / total
    center = (p + z * z / (2 * total)) / denominator
    radius = z * math.sqrt(p * (1 - p) / total + z * z / (4 * total * total)) / denominator
    return [max(0.0, center - radius), min(1.0, center + radius)]


def agreement(ratings: dict[str, list[str]]) -> dict[str, float]:
    if not ratings:
        return {}
    n_raters = {len(values) for values in ratings.values()}
    if len(n_raters) != 1 or next(iter(n_raters)) < 2:
        raise ValueError("Agreement requires a fixed number of at least two raters")
    per_item = []
    unanimous = 0
    totals = Counter()
    observed_disagreements = 0
    observed_pairs = 0
    for values in ratings.values():
        counts = Counter(values)
        n = len(values)
        agreeing = sum(count * (count - 1) for count in counts.values())
        per_item.append(agreeing / (n * (n - 1)))
        unanimous += len(counts) == 1
        totals.update(values)
        observed_disagreements += n * (n - 1) - agreeing
        observed_pairs += n * (n - 1)
    n_total = sum(totals.values())
    proportions = [count / n_total for count in totals.values()]
    expected_fleiss = sum(value * value for value in proportions)
    observed_agreement = sum(per_item) / len(per_item)
    fleiss = (
        (observed_agreement - expected_fleiss) / (1 - expected_fleiss)
        if expected_fleiss < 1
        else 1.0
    )
    observed_disagreement = observed_disagreements / observed_pairs
    expected_disagreement = sum(
        count * (n_total - count) for count in totals.values()
    ) / (n_total * (n_total - 1))
    alpha = 1 - observed_disagreement / expected_disagreement if expected_disagreement else 1.0
    return {
        "semantic_pairwise_agreement": observed_agreement,
        "semantic_unanimous_rate": unanimous / len(ratings),
        "fleiss_kappa": fleiss,
        "krippendorff_alpha_nominal": alpha,
    }


def majority(values: list[str]) -> str | None:
    value, count = Counter(values).most_common(1)[0]
    return value if count >= 2 else None


def summarize_group(
    item_ids: list[str],
    semantic_ratings: dict[str, list[str]],
    raw_responses: dict[str, list[str]],
    keys: dict[str, dict[str, str]],
) -> dict[str, Any]:
    subset = {item_id: semantic_ratings[item_id] for item_id in item_ids}
    stats: dict[str, Any] = agreement(subset)
    majority_correct = 0
    no_majority = 0
    majority_clarify = 0
    any_clarify = 0
    unanimous_correct = 0
    unanimous_items = 0
    exact_pairwise_total = 0.0
    for item_id in item_ids:
        responses = raw_responses[item_id]
        response_counts = Counter(responses)
        exact_pairwise_total += sum(
            count * (count - 1) for count in response_counts.values()
        ) / (len(responses) * (len(responses) - 1))
        vote = majority(responses)
        if vote is None:
            no_majority += 1
        else:
            majority_correct += vote == normalize_response(keys[item_id]["gold_target"])
            majority_clarify += vote == CLARIFY
        any_clarify += CLARIFY in responses
        if len(set(responses)) == 1:
            unanimous_items += 1
            unanimous_correct += responses[0] == normalize_response(keys[item_id]["gold_target"])
    total = len(item_ids)
    stats.update({
        "n_items": total,
        "exact_pairwise_agreement": exact_pairwise_total / total,
        "exact_unanimous_rate": unanimous_items / total,
        "majority_gold_correct": majority_correct,
        "majority_gold_accuracy": majority_correct / total,
        "majority_gold_wilson_95": wilson(majority_correct, total),
        "no_majority_rate": no_majority / total,
        "determinate_majority_n": total - no_majority,
        "determinate_majority_gold_accuracy": (
            majority_correct / (total - no_majority) if total > no_majority else float("nan")
        ),
        "majority_clarify_rate": majority_clarify / total,
        "any_clarify_rate": any_clarify / total,
        "consensus_only_n": unanimous_items,
        "consensus_only_gold_accuracy": (
            unanimous_correct / unanimous_items if unanimous_items else float("nan")
        ),
    })
    return stats


def payload_signature(row: dict[str, str]) -> tuple[str, ...]:
    return tuple(
        row.get(name, "").strip()
        for name in ("instruction", "initial_state", "refreshed_state", "action_schema", "candidate_ids")
    )


def markdown_report(report: dict[str, Any]) -> str:
    lines = [
        "# TRI Human Construct Validation",
        "",
        "All statistics use three independent annotations per item. Confidence was optional.",
        "",
        "## Quality audit",
        "",
    ]
    quality = report["quality"]
    for key in (
        "n_keyed_items", "n_annotators", "missing_responses", "invalid_responses",
        "duplicate_ids", "payload_mismatches", "three_way_ties",
    ):
        lines.append(f"- {key}: {quality[key]}")
    lines.extend([
        "",
        "## Agreement and gold alignment",
        "",
        "| Slice | n | Majority-gold | Determinate majority-gold | 95% Wilson CI | Unanimous | Fleiss kappa | Krippendorff alpha | Any clarify |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|",
    ])
    for name, stats in report["groups"].items():
        low, high = stats["majority_gold_wilson_95"]
        lines.append(
            f"| {name} | {stats['n_items']} | {stats['majority_gold_accuracy']:.1%} | "
            f"{stats['determinate_majority_gold_accuracy']:.1%} | [{low:.1%}, {high:.1%}] | {stats['exact_unanimous_rate']:.1%} | "
            f"{stats['fleiss_kappa']:.3f} | {stats['krippendorff_alpha_nominal']:.3f} | "
            f"{stats['any_clarify_rate']:.1%} |"
        )
    lines.extend([
        "",
        "## Annotator-level audit",
        "",
        "| Annotator | Gold accuracy | Clarify rate | Confidence coverage | Mean confidence |",
        "|---|---:|---:|---:|---:|",
    ])
    for row in report["annotators"]:
        mean = "NA" if row["mean_confidence"] is None else f"{row['mean_confidence']:.2f}"
        lines.append(
            f"| {row['annotator']} | {row['gold_accuracy']:.1%} | {row['clarify_rate']:.1%} | "
            f"{row['confidence_coverage']:.1%} | {mean} |"
        )
    lines.extend([
        "",
        "## Disagreement items",
        "",
        "| Item | Gold | A1 | A2 | A3 | Majority |",
        "|---|---|---|---|---|---|",
    ])
    for row in report["disagreements"]:
        lines.append(
            f"| {row['item_id']} | {row['gold']} | {row['responses'][0]} | "
            f"{row['responses'][1]} | {row['responses'][2]} | {row['majority'] or 'NONE'} |"
        )
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--forms", nargs=3, required=True)
    parser.add_argument("--key", required=True)
    parser.add_argument("--output", default="human_validation/analysis.json")
    parser.add_argument("--markdown-output", default="human_validation/analysis.md")
    parser.add_argument("--normalized-dir")
    args = parser.parse_args()

    key_rows = load_csv(Path(args.key))
    keys = {row["item_id"].strip(): row for row in key_rows}
    if len(keys) != len(key_rows):
        raise ValueError("Duplicate item IDs in private key")

    raw_responses: dict[str, list[str]] = defaultdict(list)
    semantic_ratings: dict[str, list[str]] = defaultdict(list)
    payloads: dict[str, list[tuple[str, ...]]] = defaultdict(list)
    normalized_forms: list[list[dict[str, str]]] = []
    duplicate_ids = 0
    missing_responses = 0
    invalid_responses = 0
    annotator_reports = []

    for annotator_index, form_path in enumerate(args.forms, start=1):
        rows = load_form(Path(form_path))
        seen: set[str] = set()
        normalized_rows = []
        correct = 0
        clarify = 0
        confidences = []
        for row in rows:
            item_id = row.get("item_id", "").strip()
            if item_id in seen:
                duplicate_ids += 1
            seen.add(item_id)
            if item_id not in keys:
                raise ValueError(f"Unknown item ID {item_id} in {form_path}")
            response = normalize_response(row.get("response", ""))
            if not response:
                missing_responses += 1
            candidates = parse_candidates(row.get("candidate_ids", ""))
            if response and response not in candidates | {REJECT, CLARIFY}:
                invalid_responses += 1
            confidence_text = row.get("confidence_1_to_5", "").strip()
            confidence = None
            if confidence_text:
                confidence = float(confidence_text)
                if confidence < 1 or confidence > 5:
                    raise ValueError(f"Invalid confidence {confidence} for {item_id}")
                confidences.append(confidence)
            raw_responses[item_id].append(response)
            semantic_ratings[item_id].append(semantic_label(response, keys[item_id]))
            payloads[item_id].append(payload_signature(row))
            correct += response == normalize_response(keys[item_id]["gold_target"])
            clarify += response == CLARIFY
            normalized_rows.append({
                "item_id": item_id,
                "response": response,
                "confidence_1_to_5": "" if confidence is None else str(confidence),
                "comment": row.get("comment", "").strip(),
            })
        if seen != set(keys):
            missing = sorted(set(keys) - seen)
            extra = sorted(seen - set(keys))
            raise ValueError(f"Form {form_path} has mismatched IDs: missing={missing}, extra={extra}")
        normalized_forms.append(normalized_rows)
        annotator_reports.append({
            "annotator": f"A{annotator_index}",
            "n_items": len(rows),
            "gold_accuracy": correct / len(rows),
            "clarify_rate": clarify / len(rows),
            "confidence_coverage": len(confidences) / len(rows),
            "mean_confidence": sum(confidences) / len(confidences) if confidences else None,
        })

    payload_mismatches = sum(len(set(values)) != 1 for values in payloads.values())
    if missing_responses or invalid_responses or duplicate_ids or payload_mismatches:
        raise ValueError(
            "Quality audit failed: "
            f"missing={missing_responses}, invalid={invalid_responses}, duplicates={duplicate_ids}, "
            f"payload_mismatches={payload_mismatches}"
        )

    groups: dict[str, list[str]] = {
        "all": sorted(keys),
        "explicit": [item_id for item_id, key in keys.items() if key["explicitness"] == "explicit"],
        "implicit": [item_id for item_id, key in keys.items() if key["explicitness"] == "implicit"],
        "original": [item_id for item_id, key in keys.items() if key["variant"] == "original"],
        "human_rewrite": [item_id for item_id, key in keys.items() if key["variant"] == "human_rewrite"],
        "anchored": [item_id for item_id, key in keys.items() if key["binding"] == "anchored"],
        "dynamic": [item_id for item_id, key in keys.items() if key["binding"] == "dynamic"],
        "anchored_actionable": [
            item_id
            for item_id, key in keys.items()
            if key["binding"] == "anchored" and normalize_response(key["gold_target"]) != REJECT
        ],
        "anchored_reject": [
            item_id
            for item_id, key in keys.items()
            if key["binding"] == "anchored" and normalize_response(key["gold_target"]) == REJECT
        ],
    }
    for update in sorted({key["update"] for key in keys.values()}):
        groups[f"update:{update}"] = [
            item_id for item_id, key in keys.items() if key["update"] == update
        ]
    for label in sorted({semantic_label(normalize_response(key["gold_target"]), key) for key in keys.values()}):
        groups[f"gold:{label}"] = [
            item_id
            for item_id, key in keys.items()
            if semantic_label(normalize_response(key["gold_target"]), key) == label
        ]

    disagreements = []
    for item_id in sorted(keys):
        responses = raw_responses[item_id]
        if len(set(responses)) > 1:
            disagreements.append({
                "item_id": item_id,
                "gold": normalize_response(keys[item_id]["gold_target"]),
                "responses": responses,
                "majority": majority(responses),
                "binding": keys[item_id]["binding"],
                "explicitness": keys[item_id]["explicitness"],
                "variant": keys[item_id]["variant"],
                "update": keys[item_id]["update"],
            })

    report = {
        "quality": {
            "n_keyed_items": len(keys),
            "n_annotators": 3,
            "missing_responses": missing_responses,
            "invalid_responses": invalid_responses,
            "duplicate_ids": duplicate_ids,
            "payload_mismatches": payload_mismatches,
            "three_way_ties": sum(majority(values) is None for values in raw_responses.values()),
            "identical_response_vectors": [
                [f"A{i + 1}", f"A{j + 1}"]
                for i in range(3)
                for j in range(i + 1, 3)
                if all(raw_responses[item_id][i] == raw_responses[item_id][j] for item_id in keys)
            ],
        },
        "groups": {
            name: summarize_group(item_ids, semantic_ratings, raw_responses, keys)
            for name, item_ids in groups.items()
        },
        "annotators": annotator_reports,
        "disagreements": disagreements,
    }

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    markdown = Path(args.markdown_output)
    markdown.parent.mkdir(parents=True, exist_ok=True)
    markdown.write_text(markdown_report(report), encoding="utf-8")

    if args.normalized_dir:
        normalized_dir = Path(args.normalized_dir)
        normalized_dir.mkdir(parents=True, exist_ok=True)
        fieldnames = ["item_id", "response", "confidence_1_to_5", "comment"]
        for index, rows in enumerate(normalized_forms, start=1):
            with (normalized_dir / f"annotator_{index}.csv").open(
                "w", encoding="utf-8-sig", newline=""
            ) as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)

    print(json.dumps(report, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
