from __future__ import annotations

import csv
import hashlib
import json
import random
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


SEED = 20260727
WRITERS = tuple(f"W{index}" for index in range(1, 13))
ANNOTATORS = tuple(f"A{index}" for index in range(1, 4))
DOMAINS = 10
PAIRS_PER_DOMAIN = 6
PAIRS = DOMAINS * PAIRS_PER_DOMAIN
INSTRUCTIONS = PAIRS * 2
EVIDENCE_STATUS = "planned/unverified"
INTENDED_STATUS = "post-primary independently authored controlled-language holdout"


def canonical_json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, ensure_ascii=True, separators=(",", ":"))


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_path(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def read_table(path: Path) -> list[dict[str, str]]:
    """Read a WJX CSV/XLSX export without retaining platform-specific value types."""
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        values = workbook.active.iter_rows(values_only=True)
        headers = ["" if value is None else str(value).strip() for value in next(values, ())]
        return [
            {
                header: "" if value is None else str(value).strip()
                for header, value in zip(headers, row)
            }
            for row in values
        ]
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [
            {key: (value or "").strip() for key, value in row.items()}
            for row in csv.DictReader(handle)
        ]


def jsonl_bytes(rows: Iterable[dict[str, Any]]) -> bytes:
    return ("\n".join(canonical_json(row) for row in rows) + "\n").encode("utf-8")


def _pair_signature(row: dict[str, Any]) -> tuple[str, str, str]:
    return row["domain"], row["state_cluster_id"], row["update"]


def build_scenario_pairs(source: Path) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in load_jsonl(source):
        if row["update"] != "stable":
            grouped[_pair_signature(row)].append(row)

    candidates: dict[str, list[list[dict[str, Any]]]] = defaultdict(list)
    for signature, rows in grouped.items():
        if len(rows) != 2 or {row["binding"] for row in rows} != {"anchored", "dynamic"}:
            raise ValueError(f"source group is not an opposite-gold pair: {signature}")
        first, second = rows
        for field in (
            "initial_state",
            "refreshed_state",
            "selector",
            "action",
            "action_schema",
            "pre_refresh_target",
            "post_refresh_target",
        ):
            if first[field] != second[field]:
                raise ValueError(f"pair field differs for {signature}: {field}")
        if first["pre_refresh_target"] == first["post_refresh_target"]:
            raise ValueError(f"pair does not change winner: {signature}")
        candidates[signature[0]].append(rows)

    if len(candidates) != DOMAINS:
        raise ValueError(f"expected {DOMAINS} domains, found {len(candidates)}")
    pairs = []
    for domain in sorted(candidates):
        domain_pairs = sorted(candidates[domain], key=lambda rows: _pair_signature(rows[0]))
        if len(domain_pairs) < PAIRS_PER_DOMAIN:
            raise ValueError(f"domain {domain} has fewer than six changed pairs")
        for local_index, rows in enumerate(domain_pairs[:PAIRS_PER_DOMAIN], start=1):
            anchored = next(row for row in rows if row["binding"] == "anchored")
            pairs.append(
                {
                    "pair_id": f"IH-{domain.upper()}-{local_index:02d}",
                    "domain": domain,
                    "source_pair": "::".join(_pair_signature(anchored)),
                    "app": anchored["app"],
                    "entity": anchored["entity"],
                    "selector": anchored["selector"],
                    "action": anchored["action"],
                    "action_schema": anchored["action_schema"],
                    "initial_state": anchored["initial_state"],
                    "refreshed_state": anchored["refreshed_state"],
                    "pre_refresh_target": anchored["pre_refresh_target"],
                    "post_refresh_target": anchored["post_refresh_target"],
                    "update": anchored["update"],
                }
            )
    validate_pairs(pairs)
    return pairs


def validate_pairs(pairs: list[dict[str, Any]]) -> None:
    if len(pairs) != PAIRS or len({row["pair_id"] for row in pairs}) != PAIRS:
        raise ValueError("holdout must contain 60 unique pairs")
    counts = Counter(row["domain"] for row in pairs)
    if len(counts) != DOMAINS or set(counts.values()) != {PAIRS_PER_DOMAIN}:
        raise ValueError("each of ten domains must contribute six pairs")
    for row in pairs:
        initial_ids = {str(entity["id"]) for entity in row["initial_state"]}
        refreshed_ids = {str(entity["id"]) for entity in row["refreshed_state"]}
        if row["pre_refresh_target"] not in initial_ids:
            raise ValueError(f"missing pre-refresh target in {row['pair_id']}")
        if row["pre_refresh_target"] not in refreshed_ids:
            raise ValueError(f"old target does not survive in {row['pair_id']}")
        if row["post_refresh_target"] not in refreshed_ids:
            raise ValueError(f"missing post-refresh target in {row['pair_id']}")


def build_assignments(
    pairs: list[dict[str, Any]],
    writers: tuple[str, ...] = WRITERS,
) -> list[dict[str, Any]]:
    if not writers or len(writers) % 2 or PAIRS % len(writers):
        raise ValueError("writers must be an even divisor of the 60 pairs")
    assignments = []
    for pair_index, pair in enumerate(pairs):
        writer_preserve = writers[pair_index % len(writers)]
        writer_reevaluate = writers[(pair_index + len(writers) // 2) % len(writers)]
        for mode, writer in (("preserve", writer_preserve), ("reevaluate", writer_reevaluate)):
            assignments.append(
                {
                    "item_id": f"{pair['pair_id']}-{'P' if mode == 'preserve' else 'R'}",
                    "pair_id": pair["pair_id"],
                    "writer_id": writer,
                    "mode": mode,
                    "domain": pair["domain"],
                }
            )
    validate_assignments(assignments, writers)
    return assignments


def validate_assignments(
    rows: list[dict[str, Any]],
    writers: tuple[str, ...] = WRITERS,
) -> None:
    if len(rows) != INSTRUCTIONS or len({row["item_id"] for row in rows}) != INSTRUCTIONS:
        raise ValueError("assignment must contain 120 unique items")
    by_pair: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_pair[row["pair_id"]].append(row)
    for pair_id, members in by_pair.items():
        if len(members) != 2 or {row["mode"] for row in members} != {"preserve", "reevaluate"}:
            raise ValueError(f"incomplete assignment pair: {pair_id}")
        if len({row["writer_id"] for row in members}) != 2:
            raise ValueError(f"pair members share a writer: {pair_id}")
    expected_per_mode = PAIRS // len(writers)
    for writer in writers:
        writer_rows = [row for row in rows if row["writer_id"] == writer]
        if Counter(row["mode"] for row in writer_rows) != Counter(
            {"preserve": expected_per_mode, "reevaluate": expected_per_mode}
        ):
            raise ValueError(f"writer {writer} is not mode-balanced")


def load_assignments(path: Path) -> list[dict[str, str]]:
    rows = read_table(path)
    required = {"item_id", "pair_id", "writer_id", "mode", "domain"}
    if not rows or not required.issubset(rows[0]):
        missing = required - set(rows[0] if rows else {})
        raise ValueError(f"writer allocation is missing columns: {sorted(missing)}")
    normalized = [{field: row[field].strip() for field in required} for row in rows]
    validate_assignments(normalized)
    return normalized


def writer_item_order(
    assignments: list[dict[str, Any]], writer_id: str
) -> list[dict[str, Any]]:
    rows = [row for row in assignments if row["writer_id"] == writer_id]
    if len(rows) != 10:
        raise ValueError(f"writer {writer_id} must have exactly 10 assigned items")
    random.Random(SEED + int(writer_id[1:])).shuffle(rows)
    return rows


def _state_text(records: list[dict[str, Any]]) -> str:
    return "\n".join(canonical_json(record) for record in records)


def _readable_state(records: list[dict[str, Any]]) -> list[str]:
    field_zh = {
        "actionable": "可执行", "display": "名称", "owner": "负责人",
        "score": "评分", "stage": "阶段", "rating": "评分", "resolved": "已解决",
        "age_minutes": "等待分钟", "status": "状态", "amount_cents": "金额",
        "severity": "严重度", "state": "状态", "value_cents": "金额",
        "start_minutes": "开始分钟", "due_minutes": "到期分钟",
        "expires_minutes": "到期分钟", "completed": "已完成",
    }
    value_zh = {
        True: "是", False: "否", "pending": "待处理", "archived": "已归档",
        "queued": "排队中", "completed": "已完成", "unpaid": "未支付",
        "paid": "已支付", "open": "开放", "closed": "关闭",
        "withdrawn": "已撤回", "unconfirmed": "未确认", "cancelled": "已取消",
        "active": "有效", "expired": "已过期",
    }
    lines = []
    for record in records:
        fields = [f"编号={record['id']}"]
        for key, value in record.items():
            if key != "id":
                fields.append(f"{field_zh.get(key, key)}={value_zh.get(value, value)}")
        lines.append("- " + " | ".join(fields))
    return lines


def _order_requirement(mode: str) -> str:
    if mode == "preserve":
        return "1）先从当前列表确定一个对象；2）同步一次；3）对刚才确定的同一个对象执行操作。"
    return "1）先同步一次；2）再从同步后的列表确定一个对象；3）对该对象执行操作。"


APP_ZH = {
    "alert_console": "告警控制台", "application_portal": "申请管理平台",
    "appointment_queue": "预约队列", "batch_manager": "批任务管理器",
    "bid_board": "投标管理板", "billing_console": "账单控制台",
    "certificate_registry": "证书登记系统", "reminder_service": "提醒事项服务",
    "reservation_board": "预订管理板", "review_inbox": "评价收件箱",
}
ACTION_ZH = {
    "acknowledge": "确认处理", "review": "审核", "confirm": "确认", "start": "启动",
    "accept": "接受", "approve": "批准", "renew": "续期", "postpone": "推迟",
    "respond": "回复", "respond to": "回复",
}
ENTITY_ZH = {
    "alert": "告警", "application": "申请", "appointment": "预约", "batch": "批任务",
    "bid": "投标", "invoice": "账单", "certificate": "证书", "reminder": "提醒事项",
    "reservation": "预订", "review": "评价",
}
SELECTOR_ZH = {
    "the highest-severity open alert": "严重度最高的开放告警",
    "the highest-scoring pending application": "评分最高的待处理申请",
    "the earliest pending appointment": "最早的待处理预约",
    "the oldest queued batch": "排队时间最长的批任务",
    "the highest-value open bid": "金额最高的开放投标",
    "the largest unpaid invoice": "金额最大的未支付账单",
    "the soonest-expiring active certificate": "最早到期的有效证书",
    "the earliest incomplete reminder": "最早到期的未完成提醒事项",
    "the earliest unconfirmed reservation": "最早的未确认预订",
    "the lowest-rated unresolved review": "评分最低的未解决评价",
}


def writer_stage_a_markdown(
    writer_id: str,
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> str:
    rows = [row for row in assignments if row["writer_id"] == writer_id]
    random.Random(SEED + int(writer_id[1:])).shuffle(rows)
    lines = [
        f"# Independent instruction writing form {writer_id}",
        "",
        "Write one natural English request for each card. Do not use examples, external tools, or",
        "another person. Submit the request before seeing the synchronized state. There is no required",
        "wording. Do not add IDs that a normal user would not know from the visible list.",
        "",
    ]
    for position, assignment in enumerate(rows, start=1):
        pair = pair_map[assignment["pair_id"]]
        lines.extend(
            [
                f"## Card {position:02d}: {assignment['item_id']}",
                "",
                f"Application: {pair['app']}",
                f"Available operation: synchronize once, then {pair['action']} exactly one {pair['entity']}.",
                f"Selection criterion: {pair['selector']}.",
                "Current visible records:",
                "BEGIN RECORDS",
                _state_text(pair["initial_state"]),
                "END RECORDS",
                f"Required workflow order: {_order_requirement(assignment['mode'])}",
                "",
                "Your single English request:",
                "",
                "____________________________________________________________",
                "",
            ]
        )
    return "\n".join(lines)


def writer_stage_b_markdown(
    writer_id: str,
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> str:
    rows = sorted(
        (row for row in assignments if row["writer_id"] == writer_id),
        key=lambda row: row["item_id"],
    )
    lines = [
        f"# Writer-intent form {writer_id}",
        "",
        "Complete this only after Stage A has been submitted. Do not revise the original request.",
        "For each item, record the saved Stage A instruction hash and then choose the intended target",
        "after viewing the synchronized state. Choose CLARIFY if the original message does not determine",
        "one target.",
        "",
    ]
    for assignment in rows:
        pair = pair_map[assignment["pair_id"]]
        candidates = sorted({str(row["id"]) for row in pair["refreshed_state"]})
        lines.extend(
            [
                f"## {assignment['item_id']}",
                "",
                "Synchronized records:",
                "BEGIN RECORDS",
                _state_text(pair["refreshed_state"]),
                "END RECORDS",
                f"Allowed intent response: {' | '.join(candidates)} | CLARIFY",
                "Stage A instruction SHA-256:",
                "Writer intent:",
                "Confidence (1-5):",
                "Optional reason:",
                "",
            ]
        )
    return "\n".join(lines)


def build_annotation_order(assignments: list[dict[str, Any]], annotator: str) -> list[str]:
    rng = random.Random(SEED + 100 + int(annotator[1:]))
    remaining = [row["item_id"] for row in assignments]
    rng.shuffle(remaining)
    order = []
    pair_by_item = {row["item_id"]: row["pair_id"] for row in assignments}
    while remaining:
        choice_index = next(
            (
                index
                for index, item_id in enumerate(remaining)
                if not order or pair_by_item[item_id] != pair_by_item[order[-1]]
            ),
            None,
        )
        if choice_index is None:
            raise ValueError("could not separate paired annotation items")
        order.append(remaining.pop(choice_index))
    return order


def _wjx_intro(stage: str) -> list[str]:
    participation = "\u2028".join(
        [
            "【参与说明】",
            "本研究了解人们如何为工具型助手撰写和理解任务请求。",
            "参与完全自愿；答案将去标识化，补偿不取决于是否符合任何预期答案。",
            "作答时请勿使用生成式 AI、搜索引擎、机器翻译或他人协助。",
        ]
    )
    return [
        f"独立语言指令研究 - {stage}",
        "",
        participation + " [段落说明]",
        "",
        "1. 你是否已年满18周岁？",
        "A. 是",
        "B. 否",
        "",
        "2. 你是否能够独立阅读情境，并写出一句简短英文请求？",
        "A. 是",
        "B. 否",
        "",
        "3. 你是否自愿参加，并同意去标识化答案用于科研？",
        "A. 是，我自愿同意",
        "B. 否",
        "",
    ]


def writer_stage_a_wjx(
    writer_id: str,
    batch: int,
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> str:
    rows = [row for row in assignments if row["writer_id"] == writer_id]
    random.Random(SEED + int(writer_id[1:])).shuffle(rows)
    rows = rows[(batch - 1) * 10 : batch * 10]
    lines = _wjx_intro(f"写作者 {writer_id}，A阶段，第 {batch}/2 部分")
    instructions = "\u2028".join(
        [
            "【作答说明】",
            "每个情境相互独立。",
            "请按给定操作顺序，用你自己的表达写一句自然的英文请求。",
            "没有固定句式，也不要照搬操作说明。",
            "提交本问卷后才会看到同步后的状态。",
        ]
    )
    lines.extend(
        [
            instructions + " [段落说明]",
            "",
        ]
    )
    for index, assignment in enumerate(rows, start=1):
        pair = pair_map[assignment["pair_id"]]
        block = [
            f"【情境 {index:02d}｜{assignment['item_id']}】",
            f"应用：{APP_ZH.get(pair['app'], pair['app'])}",
            f"工具能力：可同步一次；随后可对一个{ENTITY_ZH.get(pair['entity'], pair['entity'])}执行“{ACTION_ZH.get(pair['action'], pair['action'])}”。",
            f"筛选条件：{SELECTOR_ZH.get(pair['selector'], pair['selector'])}。",
            "当前可见记录：",
            *_readable_state(pair["initial_state"]),
            "必须表达的操作顺序：",
            _order_requirement(assignment["mode"]),
        ]
        lines.extend(
            [
                "\u2028".join(block) + " [段落说明]",
                "",
                f"{index + 3}. 请为 {assignment['item_id']} 写一句英文请求（按上面的顺序，用你自己的表达）：[填空题]",
                "",
            ]
        )
    return "\n".join(lines)


def writer_stage_b_wjx(
    writer_id: str,
    batch: int,
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> str:
    rows = sorted(
        (row for row in assignments if row["writer_id"] == writer_id),
        key=lambda row: row["item_id"],
    )[(batch - 1) * 10 : batch * 10]
    lines = _wjx_intro(f"写作者 {writer_id}，B阶段，第 {batch}/2 部分")
    instructions = "\u2028".join(
        [
            "【作答说明】",
            "仅在提交对应 A 阶段问卷后填写。",
            "不得修改先前写下的英文请求。",
            "查看同步后的记录后，选择你原来打算让助手操作的对象。",
            "若原请求无法确定唯一对象，请选“需要澄清”。",
        ]
    )
    lines.extend(
        [
            instructions + " [段落说明]",
            "",
        ]
    )
    question = 4
    for index, assignment in enumerate(rows, start=1):
        pair = pair_map[assignment["pair_id"]]
        candidates = sorted({str(row["id"]) for row in pair["refreshed_state"]})
        block = [
            f"【意图判断 {index:02d}｜{assignment['item_id']}】",
            "同步后的记录：",
            *_readable_state(pair["refreshed_state"]),
        ]
        lines.extend(["\u2028".join(block) + " [段落说明]", ""])
        lines.append(f"{question}. 你在已提交的英文请求中打算操作哪个对象？")
        lines.extend([f"{chr(65 + i)}. {target if target != 'CLARIFY' else '需要澄清'}" for i, target in enumerate(candidates + ["CLARIFY"])])
        lines.append("")
        question += 1
        lines.extend(
            [
                f"{question}. 你对这个意图判断有多确定？",
                "A. 1 - 非常不确定",
                "B. 2",
                "C. 3",
                "D. 4",
                "E. 5 - 非常确定",
                "",
            ]
        )
        question += 1
    return "\n".join(lines)


def _compact_state(records: list[dict[str, Any]]) -> str:
    field_zh = {
        "state": "状态", "stage": "阶段", "status": "状态", "resolved": "已解决",
        "completed": "已完成", "severity": "严重度", "score": "评分",
        "rating": "评分", "age_minutes": "等待分钟", "value_cents": "金额",
        "amount_cents": "金额", "start_minutes": "开始分钟",
        "due_minutes": "到期分钟", "days_remaining": "剩余天数",
    }
    value_zh = {
        True: "是", False: "否", "pending": "待处理", "archived": "已归档",
        "queued": "排队中", "completed": "已完成", "unpaid": "未支付",
        "paid": "已支付", "open": "开放", "closed": "关闭",
        "withdrawn": "已撤回", "unconfirmed": "未确认", "cancelled": "已取消",
        "active": "有效", "expired": "已过期", "revoked": "已撤销",
    }
    field_order = (
        "state", "stage", "status", "resolved", "completed", "severity", "score",
        "rating", "age_minutes", "value_cents", "amount_cents", "start_minutes",
        "due_minutes", "days_remaining",
    )
    markers = "①②③④⑤⑥⑦⑧⑨"
    rendered = []
    for index, record in enumerate(records):
        fields = [f"{markers[index]} {record['id']}", str(record.get("display", ""))]
        for key in field_order:
            if key in record:
                if key == "resolved":
                    fields.append("已解决" if record[key] else "未解决")
                elif key == "completed":
                    fields.append("已完成" if record[key] else "未完成")
                else:
                    fields.append(f"{field_zh[key]} {value_zh.get(record[key], record[key])}")
        rendered.append("｜".join(field for field in fields if field))
    return "　　".join(rendered)


def _compact_order_requirement(mode: str) -> str:
    if mode == "preserve":
        return "先按规则选对象 → 同步一次 → 操作刚才选中的同一对象"
    return "先同步一次 → 再按规则选对象 → 操作该对象"


def _wjx_paragraph(text: str) -> list[str]:
    return [text + " [段落说明]", ""]


def writer_combined_wjx(
    writer_id: str,
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
    *,
    page_size: int = 5,
    title_suffix: str = "易读回显版",
) -> str:
    """Build one paginated WJX form containing both writer stages."""
    stage_a_rows = [row for row in assignments if row["writer_id"] == writer_id]
    if not stage_a_rows or page_size <= 0 or len(stage_a_rows) % page_size:
        raise ValueError("writer items must divide evenly across positive page_size")
    random.Random(SEED + int(writer_id[1:])).shuffle(stage_a_rows)
    stage_a_question_by_item = {
        row["item_id"]: index + 4 for index, row in enumerate(stage_a_rows)
    }
    stage_b_rows = sorted(
        (row for row in assignments if row["writer_id"] == writer_id),
        key=lambda row: row["item_id"],
    )
    task_pages = len(stage_a_rows) // page_size
    total_pages = task_pages * 2
    lines = _wjx_intro(f"写作者 {writer_id}，完整问卷（{title_suffix}）")
    lines.extend(
        _wjx_paragraph(
            f"【填写顺序】本问卷共 {total_pages} 个任务页面。前 {task_pages} 页完成 "
            f"{len(stage_a_rows)} 条英文请求；后 {task_pages} 页查看同步后的列表并确认原本意图。"
            "进入 B 阶段后，请勿返回修改英文请求。"
        )
    )
    lines.extend(
        _wjx_paragraph(
            "【英文要求】每题只需一句简短英文，通常 8-25 个词；语法不必复杂，意思和顺序清楚即可。"
        )
    )
    lines.extend(
        _wjx_paragraph(
            "【格式示例】情境：先查看今天的天气，再发送提醒。可填写："
            "Check today's weather, then send me a reminder.（示例与正式任务无关，请勿照抄。）"
        )
    )

    question = 4
    for page in range(1, task_pages + 1):
        if page > 1:
            lines.extend(["[分页栏]", ""])
        lines.extend(
            _wjx_paragraph(
                f"【A 阶段｜第 {page}/{task_pages} 页】本页 {page_size} 个独立情境。"
                "请按给定顺序写一句自然英文请求。"
            )
        )
        page_rows = stage_a_rows[(page - 1) * page_size : page * page_size]
        for index, assignment in enumerate(
            page_rows, start=(page - 1) * page_size + 1
        ):
            pair = pair_map[assignment["pair_id"]]
            lines.extend(
                _wjx_paragraph(
                    f"【情境 {index:02d}｜{assignment['item_id']}】"
                    f"{APP_ZH.get(pair['app'], pair['app'])}"
                )
            )
            lines.extend(
                _wjx_paragraph(
                    f"【任务】同步一次；随后对一个{ENTITY_ZH.get(pair['entity'], pair['entity'])}执行“"
                    f"{ACTION_ZH.get(pair['action'], pair['action'])}”　｜　"
                    f"【选择规则】{SELECTOR_ZH.get(pair['selector'], pair['selector'])}"
                )
            )
            lines.extend(
                _wjx_paragraph(f"【同步前列表】{_compact_state(pair['initial_state'])}")
            )
            lines.extend(
                _wjx_paragraph(f"【必须表达的顺序】{_compact_order_requirement(assignment['mode'])}")
            )
            lines.extend(
                [
                    f"{question}. 请用一句简短、自然的英文表达上述请求（语法不必复杂）：[填空题]",
                    "",
                ]
            )
            question += 1

    for page in range(1, task_pages + 1):
        lines.extend(["[分页栏]", ""])
        lines.extend(
            _wjx_paragraph(
                f"【B 阶段｜第 {page}/{task_pages} 页】A 阶段已完成。"
                "请根据同步后的列表确认你原本想操作的对象；"
                "每题会回显你在 A 阶段写下的原句。不得返回修改英文请求。"
                "请保持原句中的操作顺序，不要因为现在看到新列表而改变原句含义。"
                "若原句无法确定唯一对象，请选“需要澄清”。"
            )
        )
        page_rows = stage_b_rows[(page - 1) * page_size : page * page_size]
        for index, assignment in enumerate(
            page_rows, start=(page - 1) * page_size + 1
        ):
            pair = pair_map[assignment["pair_id"]]
            candidates = sorted({str(row["id"]) for row in pair["refreshed_state"]})
            lines.extend(
                _wjx_paragraph(
                    f"【意图确认 {index:02d}/{len(stage_b_rows):02d}｜{assignment['item_id']}】"
                )
            )
            lines.extend(
                _wjx_paragraph(
                    f"【任务回顾】在{APP_ZH.get(pair['app'], pair['app'])}中，对一个"
                    f"{ENTITY_ZH.get(pair['entity'], pair['entity'])}执行“"
                    f"{ACTION_ZH.get(pair['action'], pair['action'])}”　｜　"
                    f"【选择规则】{SELECTOR_ZH.get(pair['selector'], pair['selector'])}"
                )
            )
            lines.extend(
                _wjx_paragraph(
                    f"【同步后列表】{_compact_state(pair['refreshed_state'])}"
                )
            )
            lines.append(
                f"{question}. 你在 A 阶段写下：“"
                f"[q{stage_a_question_by_item[assignment['item_id']]}]”。"
                "根据这条原句，你原本打算操作哪个对象？"
            )
            lines.extend(
                [
                    f"{chr(65 + i)}. {target if target != 'CLARIFY' else '需要澄清'}"
                    for i, target in enumerate(candidates + ["CLARIFY"])
                ]
            )
            lines.append("")
            question += 1
            lines.extend(
                [
                    f"{question}. 你对这个意图判断有多确定？",
                    "A. 1 - 非常不确定",
                    "B. 2",
                    "C. 3",
                    "D. 4",
                    "E. 5 - 非常确定",
                    "",
                ]
            )
            question += 1
    return "\n".join(lines)


def annotation_wjx(
    annotator: str,
    batch: int,
    item_ids: list[str],
    authored_map: dict[str, dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> str:
    lines = _wjx_intro(f"盲标者 {annotator}，第 {batch}/12 部分")
    instructions = "\u2028".join(
        [
            "【作答说明】",
            "每个情境相互独立。",
            "请判断写作者原始英文请求最终想操作哪个对象。",
            "若措辞无法确定唯一对象，请选“需要澄清”。",
            "写作者意图、配对关系、标准答案和模型输出均不会显示。",
        ]
    )
    lines.extend(
        [
            instructions + " [段落说明]",
            "",
        ]
    )
    question = 4
    for index, item_id in enumerate(item_ids, start=1):
        authored = authored_map[item_id]
        pair = pair_map[authored["pair_id"]]
        candidates = sorted({str(row["id"]) for row in pair["refreshed_state"]})
        block = [
            f"【情境 {index:02d}｜{item_id}】",
            f"写作者的英文请求：{authored['instruction']}",
            "当前可见记录：",
            *_readable_state(pair["initial_state"]),
            "同步后的记录：",
            *_readable_state(pair["refreshed_state"]),
            f"可执行操作：对一个{ENTITY_ZH.get(pair['entity'], pair['entity'])}执行“{ACTION_ZH.get(pair['action'], pair['action'])}”。",
        ]
        lines.extend(["\u2028".join(block) + " [段落说明]", ""])
        lines.append(f"{question}. 这条请求最终想操作哪个对象？")
        lines.extend([f"{chr(65 + i)}. {target if target != 'CLARIFY' else '需要澄清'}" for i, target in enumerate(candidates + ["CLARIFY"])])
        lines.append("")
        question += 1
        lines.extend(
            [
                f"{question}. 你对这个判断有多确定？",
                "A. 1 - 非常不确定",
                "B. 2",
                "C. 3",
                "D. 4",
                "E. 5 - 非常确定",
                "",
            ]
        )
        question += 1
    return "\n".join(lines)


def write_annotation_wjx_forms(
    authored: list[dict[str, Any]],
    pairs: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    output: Path,
) -> None:
    if len(authored) != INSTRUCTIONS or len({row["item_id"] for row in authored}) != INSTRUCTIONS:
        raise ValueError("annotation forms require all 120 validated writer items")
    output.mkdir(parents=True, exist_ok=True)
    authored_map = {row["item_id"]: row for row in authored}
    pair_map = {row["pair_id"]: row for row in pairs}
    order_manifest = {}
    for annotator in ANNOTATORS:
        order = build_annotation_order(assignments, annotator)
        if len(order) != INSTRUCTIONS or len(set(order)) != INSTRUCTIONS:
            raise ValueError(f"invalid annotation order for {annotator}")
        order_manifest[annotator] = order
        for batch in range(1, 13):
            item_ids = order[(batch - 1) * 10 : batch * 10]
            (output / f"annotator_{annotator}_part_{batch:02d}_wjx.txt").write_text(
                annotation_wjx(annotator, batch, item_ids, authored_map, pair_map),
                encoding="utf-8",
            )
    (output / "annotation_orders.json").write_text(
        json.dumps(order_manifest, indent=2, ensure_ascii=True) + "\n", encoding="utf-8"
    )


def _question_number(header: str) -> int | None:
    match = re.match(r"\s*(\d{1,3})\s*[.、:：]", header)
    return int(match.group(1)) if match else None


def _numbered_fields(raw: dict[str, str]) -> tuple[dict[int, str], dict[int, str]]:
    answers: dict[int, str] = {}
    headers: dict[int, str] = {}
    for header, value in raw.items():
        number = _question_number(header)
        if number is None:
            continue
        if number in answers:
            raise ValueError(f"duplicate WJX question number: {number}")
        answers[number] = str(value).strip()
        headers[number] = header
    return answers, headers


def _strip_choice(value: str) -> str:
    return re.sub(r"^\s*[A-Z]\s*[.、]\s*", "", str(value)).strip()


def _first_value(raw: dict[str, str], names: Iterable[str]) -> str:
    for name in names:
        value = str(raw.get(name, "")).strip()
        if value:
            return value
    return ""


def _yes_no(value: str, field: str) -> bool:
    normalized = _strip_choice(value).lower()
    if normalized in {"yes", "y", "true", "1", "是", "已完成", "完成"} or normalized.startswith("是，"):
        return True
    if normalized in {"no", "n", "false", "0", "否", "未完成", "没有"}:
        return False
    raise ValueError(f"missing or invalid {field} response")


def _confidence(value: str, field: str) -> int:
    normalized = _strip_choice(value)
    match = re.search(r"(?<!\d)([1-5])(?!\d)", normalized)
    if not match:
        raise ValueError(f"invalid confidence for {field}")
    return int(match.group(1))


def _target(value: str, allowed: set[str], field: str) -> str:
    normalized = _strip_choice(value)
    if "澄清" in normalized or normalized.upper() == "CLARIFY":
        return "CLARIFY"
    matches = [target for target in allowed if target in normalized]
    if len(matches) != 1:
        raise ValueError(f"invalid target for {field}")
    return matches[0]


def _submission_id(raw: dict[str, str], writer_id: str) -> str:
    value = _first_value(
        raw,
        ("response_id", "participant_id", "participant_code", "答卷编号", "序号"),
    )
    if not value:
        raise ValueError(f"missing response identifier for {writer_id}")
    return value


def normalize_wjx_writer_export(
    writer_id: str,
    raw: dict[str, str],
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """Convert one complete two-page WJX writer submission to ten locked item rows."""
    if writer_id not in WRITERS:
        raise ValueError(f"unknown writer ID: {writer_id}")
    numbered, headers = _numbered_fields(raw)
    stage_a_rows = writer_item_order(assignments, writer_id)
    stage_b_rows = sorted(stage_a_rows, key=lambda row: row["item_id"])
    if not set(range(1, 34)).issubset(numbered):
        missing = sorted(set(range(1, 34)) - set(numbered))
        raise ValueError(f"writer {writer_id} export is missing core questions: {missing}")

    age_18 = _yes_no(_first_value(raw, ("age_18",)) or numbered[1], "age_18")
    english_independent = _yes_no(
        _first_value(raw, ("english_independent",)) or numbered[2],
        "english_independent",
    )
    consent = _yes_no(_first_value(raw, ("consent",)) or numbered[3], "consent")

    no_assistance_raw = _first_value(
        raw,
        ("no_assistance", "未使用辅助确认", "独立作答确认"),
    ) or numbered.get(34, "")
    if no_assistance_raw:
        no_assistance = _yes_no(no_assistance_raw, "no_assistance")
    else:
        used_assistance_raw = _first_value(raw, ("used_assistance", "使用辅助"))
        if not used_assistance_raw:
            raise ValueError(f"missing no-assistance confirmation for {writer_id}")
        no_assistance = not _yes_no(used_assistance_raw, "used_assistance")

    technical_raw = _first_value(raw, ("technical_issue", "技术问题")) or numbered.get(35, "")
    if not technical_raw:
        raise ValueError(f"missing technical-issue response for {writer_id}")
    technical_issue = _yes_no(technical_raw, "technical_issue")

    completed_raw = _first_value(
        raw,
        ("completed", "completion_confirmed", "答题状态", "完成状态"),
    ) or numbered.get(36, "")
    completed = _yes_no(completed_raw, "completed")
    submission_id = _submission_id(raw, writer_id)

    stage_a: dict[str, dict[str, Any]] = {}
    for offset, assignment in enumerate(stage_a_rows):
        question = 4 + offset
        instruction = numbered[question].strip()
        if not instruction:
            raise ValueError(f"empty Stage A instruction for {assignment['item_id']}")
        stage_a[assignment["item_id"]] = {
            "instruction": instruction,
            "instruction_sha256": sha256_bytes(instruction.encode("utf-8")),
        }

    normalized = []
    for offset, assignment in enumerate(stage_b_rows):
        intent_question = 14 + 2 * offset
        confidence_question = intent_question + 1
        pair = pair_map[assignment["pair_id"]]
        allowed = {str(entity["id"]) for entity in pair["refreshed_state"]}
        intent = _target(numbered[intent_question], allowed, assignment["item_id"])

        # WJX may expand the dynamic reference in the exported question header. If it does,
        # require the displayed Stage A text to match the string hashed above.
        header = headers.get(intent_question, "")
        echo_match = re.search(r"写下[：:]?[“\"](.+?)[”\"]", header)
        if echo_match and "[q" not in echo_match.group(1):
            if echo_match.group(1) != stage_a[assignment["item_id"]]["instruction"]:
                raise ValueError(f"Stage A echo mismatch for {assignment['item_id']}")

        normalized.append(
            {
                **assignment,
                **stage_a[assignment["item_id"]],
                "writer_intent": intent,
                "writer_confidence": _confidence(
                    numbered[confidence_question], assignment["item_id"]
                ),
                "writer_submission_id": submission_id,
                "age_18": age_18,
                "english_independent": english_independent,
                "consent": consent,
                "no_assistance": no_assistance,
                "technical_issue": technical_issue,
                "completed": completed,
            }
        )
    return normalized


def validate_writer_returns(
    rows: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    pair_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    assigned = {row["item_id"]: row for row in assignments}
    observed_ids = [row.get("item_id") for row in rows]
    if (
        len(rows) != INSTRUCTIONS
        or len(set(observed_ids)) != INSTRUCTIONS
        or set(observed_ids) != set(assigned)
    ):
        raise ValueError("writer returns must contain all 120 assigned items exactly once")
    by_writer: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_writer[str(row.get("writer_id", ""))].append(row)
    if set(by_writer) != set(WRITERS) or any(len(group) != 10 for group in by_writer.values()):
        raise ValueError("writer returns must contain one complete 10-item form for W1-W12")
    submission_ids = []
    for writer_id, group in by_writer.items():
        ids = {str(row.get("writer_submission_id", "")).strip() for row in group}
        if len(ids) != 1 or not next(iter(ids)):
            raise ValueError(f"writer {writer_id} has inconsistent response identifiers")
        submission_ids.append(next(iter(ids)))
        for field in ("age_18", "english_independent", "consent", "no_assistance", "completed"):
            if {row.get(field) for row in group} != {True}:
                raise ValueError(f"writer {writer_id} failed eligibility field {field}")
        if {row.get("technical_issue") for row in group} != {False}:
            raise ValueError(f"writer {writer_id} reported an understanding-affecting technical issue")
    if len(set(submission_ids)) != len(WRITERS):
        raise ValueError("writer response identifiers must be unique")
    normalized = []
    for row in rows:
        assignment = assigned[row["item_id"]]
        if row.get("writer_id") != assignment["writer_id"]:
            raise ValueError(f"wrong writer for {row['item_id']}")
        instruction = row.get("instruction", "").strip()
        if not instruction:
            raise ValueError(f"empty instruction for {row['item_id']}")
        expected_hash = sha256_bytes(instruction.encode("utf-8"))
        if row.get("instruction_sha256") != expected_hash:
            raise ValueError(f"instruction hash mismatch for {row['item_id']}")
        pair = pair_map[assignment["pair_id"]]
        intent = row.get("writer_intent", "").strip()
        allowed = {str(entity["id"]) for entity in pair["refreshed_state"]} | {"CLARIFY"}
        if intent not in allowed:
            raise ValueError(f"invalid writer intent for {row['item_id']}")
        try:
            confidence = int(row["writer_confidence"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"invalid writer confidence for {row['item_id']}") from exc
        if confidence not in range(1, 6):
            raise ValueError(f"invalid writer confidence for {row['item_id']}")
        normalized.append(
            {
                **assignment,
                "instruction": instruction,
                "instruction_sha256": expected_hash,
                "writer_intent": intent,
                "writer_intent_determinate": intent != "CLARIFY",
                "writer_confidence": confidence,
                "allowed_target_ids": sorted(allowed),
            }
        )
    return sorted(normalized, key=lambda row: row["item_id"])


def validate_annotation_returns(
    rows: list[dict[str, str]],
    authored: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    expected_items = {row["item_id"] for row in authored}
    expected = {(annotator, item_id) for annotator in ANNOTATORS for item_id in expected_items}
    observed_pairs = [(row.get("annotator_id"), row.get("item_id")) for row in rows]
    if (
        len(rows) != len(expected)
        or len(set(observed_pairs)) != len(expected)
        or set(observed_pairs) != expected
    ):
        raise ValueError("annotation returns must contain 3 labels for every item")
    if {row.get("annotator_id") for row in rows} != set(ANNOTATORS):
        raise ValueError("annotator IDs must be the three unique frozen IDs")
    authored_map = {row["item_id"]: row for row in authored}
    normalized = []
    for row in rows:
        item = authored_map[row["item_id"]]
        target = str(row.get("target", "")).strip()
        allowed = set(item.get("allowed_target_ids", ()))
        if target not in allowed:
            raise ValueError(f"invalid annotation target for {row['item_id']}")
        try:
            confidence = int(row["confidence"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"invalid annotation confidence for {row['item_id']}") from exc
        if confidence not in range(1, 6):
            raise ValueError(f"invalid annotation confidence for {row['item_id']}")
        normalized.append(
            {
                "annotator_id": row["annotator_id"],
                "item_id": row["item_id"],
                "target": target,
                "confidence": confidence,
                "matches_writer_intent": target == item["writer_intent"],
            }
        )
    return normalized


def clear_complete_pairs(
    authored: list[dict[str, Any]],
    annotations: list[dict[str, Any]],
) -> dict[str, Any]:
    by_item = defaultdict(list)
    for row in annotations:
        by_item[row["item_id"]].append(row)
    item_clear = {}
    for item in authored:
        labels = by_item[item["item_id"]]
        agreeing = sum(row["matches_writer_intent"] for row in labels)
        item_clear[item["item_id"]] = item["writer_intent_determinate"] and agreeing >= 2
    by_pair = defaultdict(list)
    for item in authored:
        by_pair[item["pair_id"]].append(item)
    clear_pairs = [
        pair_id
        for pair_id, members in by_pair.items()
        if len(members) == 2 and all(item_clear[row["item_id"]] for row in members)
    ]
    return {
        "clear_items": sum(item_clear.values()),
        "clear_complete_pairs": len(clear_pairs),
        "main_paper_threshold_met": len(clear_pairs) >= 40,
        "clear_pair_ids": sorted(clear_pairs),
        "item_clear": item_clear,
    }


def build_model_tasks(
    authored: list[dict[str, Any]],
    pairs: list[dict[str, Any]],
    clarity: dict[str, Any],
) -> list[dict[str, Any]]:
    if len(authored) != INSTRUCTIONS:
        raise ValueError("model inventory requires all 120 authored instructions")
    if not clarity.get("main_paper_threshold_met") or len(clarity.get("clear_pair_ids", ())) < 40:
        raise ValueError("model calls require at least 40 clear complete human-authored pairs")
    pair_map = {row["pair_id"]: row for row in pairs}
    tasks = []
    for item in authored:
        pair = pair_map[item["pair_id"]]
        design_target = (
            pair["pre_refresh_target"]
            if item["mode"] == "preserve"
            else pair["post_refresh_target"]
        )
        writer_target = item["writer_intent"] if item["writer_intent_determinate"] else None
        tasks.append(
            {
                "id": f"tri-independent-holdout-{item['item_id']}",
                "item_id": item["item_id"],
                "pair_id": item["pair_id"],
                "state_cluster_id": item["pair_id"],
                "domain": pair["domain"],
                "source": "independently authored controlled-language holdout",
                "writer_id": item["writer_id"],
                "reference_mode_design": item["mode"],
                "instruction": item["instruction"],
                "instruction_sha256": item["instruction_sha256"],
                "initial_state": pair["initial_state"],
                "refreshed_state": pair["refreshed_state"],
                "s0_summary": {
                    "app": pair["app"],
                    "entity_type": pair["entity"],
                    "records": pair["initial_state"],
                },
                "initial_selected_id": pair["pre_refresh_target"],
                "pre_refresh_target": pair["pre_refresh_target"],
                "post_refresh_target": pair["post_refresh_target"],
                "selector": pair["selector"],
                "action": pair["action"],
                "action_schema": pair["action_schema"],
                "design_target": design_target,
                "writer_intent": item["writer_intent"],
                "correct_target": writer_target,
                "writer_intent_determinate": item["writer_intent_determinate"],
                "clear_item": bool(clarity["item_clear"][item["item_id"]]),
                "clear_complete_pair": item["pair_id"] in set(clarity["clear_pair_ids"]),
                "actionable_core": writer_target is not None,
            }
        )
    validate_model_tasks(tasks)
    return sorted(tasks, key=lambda row: (row["pair_id"], row["item_id"]))


def validate_model_tasks(tasks: list[dict[str, Any]]) -> None:
    if len(tasks) != INSTRUCTIONS or len({row["id"] for row in tasks}) != INSTRUCTIONS:
        raise ValueError("model inventory must contain 120 unique rows")
    by_pair = defaultdict(list)
    for row in tasks:
        by_pair[row["pair_id"]].append(row)
        if row["pre_refresh_target"] == row["post_refresh_target"]:
            raise ValueError(f"model task is not changed-winner: {row['id']}")
        if row["clear_item"] and row["correct_target"] is None:
            raise ValueError(f"clear task has no determinate target: {row['id']}")
    if len(by_pair) != PAIRS:
        raise ValueError("model inventory must retain all 60 pairs")
    for pair_id, members in by_pair.items():
        if len(members) != 2:
            raise ValueError(f"incomplete model pair: {pair_id}")
        for field in (
            "initial_state",
            "refreshed_state",
            "selector",
            "action",
            "action_schema",
            "pre_refresh_target",
            "post_refresh_target",
        ):
            if members[0][field] != members[1][field]:
                raise ValueError(f"model pair field differs: {pair_id} {field}")
        if len({row["writer_id"] for row in members}) != 2:
            raise ValueError(f"model pair shares a writer: {pair_id}")


def write_packet(source: Path, output: Path) -> dict[str, Any]:
    output.mkdir(parents=True, exist_ok=True)
    pairs = build_scenario_pairs(source)
    assignments = build_assignments(pairs)
    pair_map = {row["pair_id"]: row for row in pairs}
    (output / "private_scenario_key.jsonl").write_bytes(jsonl_bytes(pairs))
    with (output / "writer_allocation.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(assignments[0]))
        writer.writeheader()
        writer.writerows(assignments)
    for writer_id in WRITERS:
        (output / f"writer_{writer_id}_two_page_wjx.txt").write_text(
            writer_combined_wjx(
                writer_id,
                assignments,
                pair_map,
                page_size=10,
                title_suffix="12人版最终",
            ),
            encoding="utf-8",
        )
    orders = {annotator: build_annotation_order(assignments, annotator) for annotator in ANNOTATORS}
    (output / "annotation_orders.json").write_text(
        json.dumps(orders, indent=2, ensure_ascii=True) + "\n", encoding="utf-8"
    )
    manifest = {
        "evidence_status": EVIDENCE_STATUS,
        "intended_status_after_completion": INTENDED_STATUS,
        "seed": SEED,
        "source": str(source),
        "source_sha256": sha256_path(source),
        "pairs": PAIRS,
        "instructions": INSTRUCTIONS,
        "domains": DOMAINS,
        "writers": len(WRITERS),
        "annotators": len(ANNOTATORS),
        "clear_pair_main_paper_threshold": 40,
        "pages_per_writer": 2,
        "items_per_writer": 10,
        "rule": "pair members use different writers; each writer has 5 Preserve and 5 Reevaluate items",
        "files": {},
    }
    for path in sorted(output.iterdir()):
        if path.name != "freeze_manifest.json":
            manifest["files"][path.name] = sha256_path(path)
    (output / "freeze_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=True) + "\n", encoding="utf-8"
    )
    return manifest
