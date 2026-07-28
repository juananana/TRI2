from __future__ import annotations

import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


FORMS = tuple("ABCDEF")
LABELS_PER_ITEM = 5
ITEMS_PER_FORM = 12
EVIDENCE_STATUS = "post-primary human construct audit"
INCOMPLETE_EVIDENCE_STATUS = "post-primary audit (failed frozen eligibility gate; descriptive boundary evidence)"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [{key: (value or "").strip() for key, value in row.items()} for row in csv.DictReader(handle)]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value).strip()


def read_xlsx(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers = [_cell_text(value) for value in next(rows, ())]
    return [
        {header: _cell_text(value) for header, value in zip(headers, row)}
        for row in rows
    ]


def read_table(path: Path) -> list[dict[str, str]]:
    return read_xlsx(path) if path.suffix.lower() == ".xlsx" else read_csv(path)


def load_key(path: Path) -> list[dict[str, str]]:
    rows = read_csv(path)
    required = {
        "public_item_id",
        "source_task_id",
        "form",
        "position",
        "category",
        "discourse_referent_gold",
        "execution_gold",
    }
    if not rows or not required.issubset(rows[0]):
        raise ValueError(f"answer key is missing columns: {sorted(required - set(rows[0] if rows else {}))}")
    if len(rows) != 72 or Counter(row["form"] for row in rows) != Counter({form: 12 for form in FORMS}):
        raise ValueError("answer key must contain 12 unique items for each of six forms")
    if len({row["public_item_id"] for row in rows}) != 72:
        raise ValueError("answer key item IDs are not unique")
    return rows


def load_allocation(path: Path) -> dict[str, dict[str, str]]:
    rows = read_csv(path)
    allocation = {row["participant_code"]: row for row in rows}
    if len(allocation) != 36:
        raise ValueError("allocation must contain 30 primary and six reserve codes")
    if Counter(row["valid_status"] for row in rows) != Counter({"primary": 30, "reserve": 6}):
        raise ValueError("allocation role counts differ from the frozen design")
    return allocation


def _question_number(header: str) -> int | None:
    match = re.match(r"\s*(\d{1,2})\s*[.、:：]", header)
    return int(match.group(1)) if match else None


def _strip_choice_prefix(value: str) -> str:
    return re.sub(r"^\s*[A-Z]\s*[.、]\s*", "", value).strip()


def _yes(value: str) -> bool:
    value = _strip_choice_prefix(value).lower()
    return value in {"yes", "y", "true", "1", "是"} or value.startswith("我已阅读")


def _no(value: str) -> bool:
    return _strip_choice_prefix(value).lower() in {"no", "n", "false", "0", "否", "没有"}


def _candidate(value: str, candidates: Iterable[str]) -> str | None:
    matches = [candidate for candidate in candidates if candidate and candidate in value]
    return matches[0] if len(matches) == 1 else None


def _referent(value: str, candidates: Iterable[str] = ()) -> str | None:
    value = _strip_choice_prefix(value)
    if not value:
        return None
    if "无法唯一确定" in value or value.upper() in {"AMBIGUOUS", "CLARIFY"}:
        return "AMBIGUOUS"
    candidate = _candidate(value, candidates)
    if candidate:
        return candidate
    match = re.search(r"(?:对象\s*)?([A-Za-z][A-Za-z0-9_-]*-\d+[A-Za-z0-9_-]*)", value)
    return match.group(1) if match else re.sub(r"^对象\s*", "", value)


def _execution(value: str, candidates: Iterable[str] = ()) -> str | None:
    value = _strip_choice_prefix(value)
    if not value:
        return None
    if "拒绝" in value or value.upper() == "REJECT":
        return "REJECT"
    if "澄清" in value or value.upper() == "CLARIFY":
        return "CLARIFY"
    candidate = _candidate(value, candidates)
    if candidate:
        return candidate
    match = re.search(r"(?:执行\s*)?([A-Za-z][A-Za-z0-9_-]*-\d+[A-Za-z0-9_-]*)", value)
    return match.group(1) if match else re.sub(r"^执行\s*", "", value)


def _confidence(value: str) -> int | None:
    value = _strip_choice_prefix(value)
    if not value:
        return None
    digits = re.findall(r"[1-5]", value)
    return int(digits[-1]) if digits else None


def normalize_export_row(
    raw: dict[str, str],
    *,
    key_rows: list[dict[str, str]],
    allocation: dict[str, dict[str, str]],
    participant_map: dict[str, str] | None = None,
) -> dict[str, Any]:
    participant_map = participant_map or {}
    participant_code = raw.get("participant_code", "").strip()
    if not participant_code:
        response_id = raw.get("response_id") or raw.get("答卷编号") or raw.get("序号") or ""
        participant_code = participant_map.get(str(response_id).strip(), "")
    if participant_code not in allocation:
        raise ValueError(f"unassigned participant code: {participant_code or '<missing>'}")
    assigned = allocation[participant_code]
    form = (raw.get("form") or assigned["form"]).strip().upper()
    if form != assigned["form"]:
        raise ValueError(f"participant {participant_code} submitted the wrong form")

    numbered = {_question_number(key): value for key, value in raw.items() if _question_number(key)}
    submitted = (
        raw.get("submitted_at")
        or raw.get("提交答卷时间")
        or raw.get("提交时间")
        or raw.get("开始时间")
        or ""
    )
    items = sorted((row for row in key_rows if row["form"] == form), key=lambda row: int(row["position"]))
    responses: dict[str, dict[str, Any]] = {}
    for offset, item in enumerate(items):
        base = 4 + 3 * offset
        item_id = item["public_item_id"]
        candidates = [part.strip() for part in item.get("candidate_order", "").split("|")]
        responses[item_id] = {
            "referent": _referent(
                raw.get(f"{item_id}_referent", "") or numbered.get(base, ""), candidates
            ),
            "execution": _execution(
                raw.get(f"{item_id}_execution", "") or numbered.get(base + 1, ""), candidates
            ),
            "confidence": _confidence(raw.get(f"{item_id}_confidence", "") or numbered.get(base + 2, "")),
        }
    return {
        "participant_code": participant_code,
        "form": form,
        "allocation_role": assigned["valid_status"],
        "submitted_at": submitted,
        "age_18": _yes(raw.get("age_18", "") or numbered.get(1, "")),
        "english_independent": _yes(raw.get("english_independent", "") or numbered.get(2, "")),
        "consent": _yes(raw.get("consent", "") or numbered.get(3, "")),
        "used_assistance": not _no(raw.get("used_assistance", "") or numbered.get(40, "")),
        "technical_issue": not _no(raw.get("technical_issue", "") or numbered.get(41, "")),
        "end_check_complete": bool(raw.get("english_difficulty", "") or numbered.get(42, "")),
        "responses": responses,
    }


def is_valid_submission(row: dict[str, Any]) -> tuple[bool, list[str]]:
    reasons = []
    for field in ("age_18", "english_independent", "consent", "end_check_complete"):
        if not row[field]:
            reasons.append(field)
    if row["used_assistance"]:
        reasons.append("used_assistance")
    if row["technical_issue"]:
        reasons.append("technical_issue")
    if any(
        answer["referent"] is None
        or answer["execution"] is None
        or answer["confidence"] is None
        for answer in row["responses"].values()
    ):
        reasons.append("incomplete_items")
    return not reasons, reasons


def _time_key(value: str) -> tuple[int, str]:
    if not value:
        return (1, "")
    try:
        return (0, datetime.fromisoformat(value.replace("/", "-")).isoformat())
    except ValueError:
        return (0, value)


def select_frozen_sample(
    rows: Iterable[dict[str, Any]], *, require_complete: bool = True
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = list(rows)
    selected: list[dict[str, Any]] = []
    ledger: list[dict[str, Any]] = []
    for form in FORMS:
        form_rows = sorted(
            (row for row in rows if row["form"] == form),
            key=lambda row: (_time_key(row["submitted_at"]), row["participant_code"]),
        )
        valid_primary = []
        valid_reserve = []
        seen_codes = set()
        for row in form_rows:
            if row["participant_code"] in seen_codes:
                raise ValueError(f"duplicate submission for {row['participant_code']}")
            seen_codes.add(row["participant_code"])
            valid, reasons = is_valid_submission(row)
            ledger.append(
                {
                    "participant_code": row["participant_code"],
                    "form": form,
                    "allocation_role": row["allocation_role"],
                    "valid": valid,
                    "exclusion_reasons": reasons,
                }
            )
            if valid and row["allocation_role"] == "primary":
                valid_primary.append(row)
            elif valid and row["allocation_role"] == "reserve":
                valid_reserve.append(row)
        chosen = valid_primary[:5]
        chosen.extend(valid_reserve[: 5 - len(chosen)])
        if require_complete and len(chosen) != 5:
            raise ValueError(f"form {form} has only {len(chosen)} valid submissions")
        selected.extend(chosen)
    selected_codes = {row["participant_code"] for row in selected}
    for entry in ledger:
        entry["selected"] = entry["participant_code"] in selected_codes
    return selected, ledger


def _majority(labels: list[str]) -> str | None:
    counts = Counter(labels)
    label, count = counts.most_common(1)[0]
    if count < math.floor(len(labels) / 2) + 1:
        return None
    if sum(value == count for value in counts.values()) > 1:
        return None
    return label


def fleiss_kappa(label_sets: list[list[str]]) -> float:
    if not label_sets or len({len(labels) for labels in label_sets}) != 1:
        raise ValueError("Fleiss kappa requires a nonempty fixed-label matrix")
    n = len(label_sets[0])
    categories = sorted({label for labels in label_sets for label in labels})
    totals = Counter(label for labels in label_sets for label in labels)
    p_bar = sum(
        sum(count * count for count in Counter(labels).values()) - n
        for labels in label_sets
    ) / (len(label_sets) * n * (n - 1))
    p_e = sum((totals[label] / (len(label_sets) * n)) ** 2 for label in categories)
    return (p_bar - p_e) / (1 - p_e) if p_e < 1 else 1.0


def krippendorff_alpha_nominal(label_sets: list[list[str]]) -> float:
    values = [label for labels in label_sets for label in labels]
    if not values:
        raise ValueError("alpha requires labels")
    observed_disagreement = sum(
        sum(a != b for i, a in enumerate(labels) for b in labels[i + 1 :])
        / (len(labels) * (len(labels) - 1) / 2)
        for labels in label_sets
    ) / len(label_sets)
    counts = Counter(values)
    total = len(values)
    expected_disagreement = 1 - sum((count / total) ** 2 for count in counts.values())
    return 1 - observed_disagreement / expected_disagreement if expected_disagreement else 1.0


def _pair_id(source_task_id: str) -> str:
    return re.sub(r"-(explicit|implicit)_(anchor|dynamic)-", r"-\1_MODE-", source_task_id)


def _semantic_role(label: str | None, item: dict[str, str]) -> str:
    if label is None:
        return "MISSING"
    if label in {"REJECT", "CLARIFY", "AMBIGUOUS"}:
        return label
    pre_target = item.get("pre_refresh_target")
    post_target = item.get("post_refresh_target")
    if label == pre_target == post_target:
        return "UNCHANGED_TARGET"
    if label == pre_target:
        return "PRE_REFRESH_TARGET"
    if label == post_target:
        return "POST_REFRESH_TARGET"
    return "OTHER_CANDIDATE"


def analyze(selected: list[dict[str, Any]], key_rows: list[dict[str, str]]) -> dict[str, Any]:
    if len(selected) != 30:
        raise ValueError("analysis requires exactly 30 selected participants")
    key = {row["public_item_id"]: row for row in key_rows}
    labels: dict[str, dict[str, list[Any]]] = {
        item_id: {"referent": [], "execution": [], "confidence": []} for item_id in key
    }
    for participant in selected:
        expected = {item_id for item_id, item in key.items() if item["form"] == participant["form"]}
        if set(participant["responses"]) != expected:
            raise ValueError(f"form {participant['form']} item set differs from the frozen key")
        for item_id, answers in participant["responses"].items():
            for field in labels[item_id]:
                labels[item_id][field].append(answers[field])
    if any(len(values[field]) != LABELS_PER_ITEM for values in labels.values() for field in values):
        raise ValueError("every item must have exactly five labels")

    item_results = []
    for item_id, item_labels in labels.items():
        item = key[item_id]
        referent_majority = _majority(item_labels["referent"])
        execution_majority = _majority(item_labels["execution"])
        item_results.append(
            {
                "public_item_id": item_id,
                "source_task_id": item["source_task_id"],
                "form": item["form"],
                "category": item["category"],
                "pair_id": _pair_id(item["source_task_id"]) if item["category"] == "changed" else None,
                "referent_gold": item["discourse_referent_gold"],
                "execution_gold": item["execution_gold"],
                "referent_labels": item_labels["referent"],
                "execution_labels": item_labels["execution"],
                "confidence_labels": item_labels["confidence"],
                "referent_majority": referent_majority,
                "execution_majority": execution_majority,
                "referent_majority_gold": referent_majority == item["discourse_referent_gold"],
                "execution_majority_gold": execution_majority == item["execution_gold"],
                "referent_unanimous": len(set(item_labels["referent"])) == 1,
                "execution_unanimous": len(set(item_labels["execution"])) == 1,
                "identity_correct_execution_disagrees": (
                    referent_majority == item["discourse_referent_gold"]
                    and execution_majority != item["execution_gold"]
                ),
            }
        )

    changed = [row for row in item_results if row["category"] == "changed"]
    by_pair: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in changed:
        by_pair[row["pair_id"]].append(row)
    if len(by_pair) != 18 or any(len(rows) != 2 for rows in by_pair.values()):
        raise ValueError("changed slice must contain 18 complete matched pairs")

    execution_counts = Counter(row["execution_majority"] or "NO_MAJORITY" for row in item_results)
    referent_role_sets = [
        [_semantic_role(label, key[row["public_item_id"]]) for label in row["referent_labels"]]
        for row in item_results
    ]
    execution_role_sets = [
        [_semantic_role(label, key[row["public_item_id"]]) for label in row["execution_labels"]]
        for row in item_results
    ]
    return {
        "evidence_status": EVIDENCE_STATUS,
        "participants": len(selected),
        "items": len(item_results),
        "labels_per_item": LABELS_PER_ITEM,
        "complete_changed_pairs": len(by_pair),
        "referent": {
            "majority_gold": sum(row["referent_majority_gold"] for row in item_results),
            "unanimous": sum(row["referent_unanimous"] for row in item_results),
            "agreement_label_space": "semantic target roles",
            "fleiss_kappa": fleiss_kappa(referent_role_sets),
            "krippendorff_alpha": krippendorff_alpha_nominal(referent_role_sets),
            "pair_majority_correct": sum(
                all(row["referent_majority_gold"] for row in rows) for rows in by_pair.values()
            ),
            "pair_denominator": len(by_pair),
        },
        "execution": {
            "majority_gold": sum(row["execution_majority_gold"] for row in item_results),
            "unanimous": sum(row["execution_unanimous"] for row in item_results),
            "agreement_label_space": "semantic target roles",
            "fleiss_kappa": fleiss_kappa(execution_role_sets),
            "krippendorff_alpha": krippendorff_alpha_nominal(execution_role_sets),
            "majority_distribution": dict(sorted(execution_counts.items())),
            "clarify_majorities": execution_counts["CLARIFY"],
            "identity_correct_execution_disagrees": sum(
                row["identity_correct_execution_disagrees"] for row in item_results
            ),
        },
        "items_detail": item_results,
    }


def _label_accuracy(
    participants: list[dict[str, Any]], key: dict[str, dict[str, str]], field: str, gold: str
) -> dict[str, Any]:
    by_category: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    correct = 0
    total = 0
    participant_rates = []
    for participant in participants:
        participant_correct = 0
        participant_total = 0
        for item_id, answer in participant["responses"].items():
            is_correct = answer[field] == key[item_id][gold]
            correct += int(is_correct)
            total += 1
            participant_correct += int(is_correct)
            participant_total += 1
            category = key[item_id]["category"]
            by_category[category][0] += int(is_correct)
            by_category[category][1] += 1
        if participant_total:
            participant_rates.append(participant_correct / participant_total)
    return {
        "correct": correct,
        "total": total,
        "rate": correct / total if total else None,
        "by_category": {
            category: {"correct": values[0], "total": values[1], "rate": values[0] / values[1]}
            for category, values in sorted(by_category.items())
        },
        "participant_rate_range": {
            "minimum": min(participant_rates) if participant_rates else None,
            "maximum": max(participant_rates) if participant_rates else None,
        },
    }


def analyze_incomplete(
    selected: list[dict[str, Any]],
    key_rows: list[dict[str, str]],
    *,
    all_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    key = {row["public_item_id"]: row for row in key_rows}
    by_form = Counter(row["form"] for row in selected)
    labels_per_item = Counter()
    item_labels: dict[str, dict[str, list[Any]]] = {
        item_id: {"referent": [], "execution": [], "confidence": []} for item_id in key
    }
    for participant in selected:
        expected = {item_id for item_id, item in key.items() if item["form"] == participant["form"]}
        if set(participant["responses"]) != expected:
            raise ValueError(f"form {participant['form']} item set differs from the frozen key")
        for item_id, answers in participant["responses"].items():
            labels_per_item[item_id] += 1
            for field in item_labels[item_id]:
                item_labels[item_id][field].append(answers[field])

    majority_subset = []
    for item_id, labels in item_labels.items():
        if len(labels["referent"]) < 3:
            continue
        majority_subset.append(
            {
                "referent": _majority(labels["referent"]),
                "execution": _majority(labels["execution"]),
                "referent_gold": key[item_id]["discourse_referent_gold"],
                "execution_gold": key[item_id]["execution_gold"],
            }
        )

    confidence = [
        answer["confidence"]
        for participant in selected
        for answer in participant["responses"].values()
    ]
    report: dict[str, Any] = {
        "evidence_status": INCOMPLETE_EVIDENCE_STATUS,
        "quality_gate": {
            "passed": False,
            "required_valid_per_form": 5,
            "selected_valid_by_form": {form: by_form[form] for form in FORMS},
            "additional_valid_needed_by_form": {form: 5 - by_form[form] for form in FORMS},
            "selected_valid_total": len(selected),
            "required_valid_total": 30,
        },
        "coverage": {
            "items": len(key),
            "labels": sum(labels_per_item.values()),
            "labels_per_item_distribution": {
                str(count): frequency
                for count, frequency in sorted(Counter(labels_per_item.values()).items())
            },
            "items_with_at_least_three_labels": len(majority_subset),
            "complete_changed_pairs_at_five_labels": 0,
        },
        "eligible_exploratory": {
            "referent": _label_accuracy(
                selected, key, "referent", "discourse_referent_gold"
            ),
            "execution": _label_accuracy(selected, key, "execution", "execution_gold"),
            "three_or_more_label_item_majority": {
                "items": len(majority_subset),
                "referent_majority_gold": sum(
                    row["referent"] == row["referent_gold"] for row in majority_subset
                ),
                "execution_majority_gold": sum(
                    row["execution"] == row["execution_gold"] for row in majority_subset
                ),
            },
            "confidence_distribution": {
                str(score): count for score, count in sorted(Counter(confidence).items())
            },
        },
    }
    if all_rows is not None:
        selected_codes = {row["participant_code"] for row in selected}
        excluded = [row for row in all_rows if row["participant_code"] not in selected_codes]
        report["excluded_sensitivity"] = {
            "participants": len(excluded),
            "referent": _label_accuracy(
                excluded, key, "referent", "discourse_referent_gold"
            ),
            "execution": _label_accuracy(excluded, key, "execution", "execution_gold"),
        }
    return report


def render_markdown(report: dict[str, Any]) -> str:
    if report["evidence_status"] != INCOMPLETE_EVIDENCE_STATUS:
        return (
            "# Six-Form Human Construct Audit\n\n"
            f"Evidence status: **{report['evidence_status']}**.\n\n"
            f"The frozen analysis includes {report['participants']} participants, "
            f"{report['items']} items, and {report['complete_changed_pairs']} complete changed pairs.\n"
        )

    gate = report["quality_gate"]
    coverage = report["coverage"]
    exploratory = report["eligible_exploratory"]
    form_counts = " | ".join(
        f"{form}: {gate['selected_valid_by_form'][form]}/5" for form in FORMS
    )
    collection = report.get("collection", {})
    patterns = collection.get("exclusion_pattern_counts", {})
    duration = collection.get("duration_seconds", {})
    needed = gate["additional_valid_needed_by_form"]
    lines = [
        "# Six-Form Human Construct Audit: Incomplete Cutoff",
        "",
        f"Evidence status: **{report['evidence_status']}**. The completed collection did not pass the frozen sample gate and cannot provide a fixed-rater construct endpoint.",
        "",
        "## Eligibility Gate",
        "",
        f"The cutoff contains {collection.get('mapped_total', 0)} complete submissions: {collection.get('valid_total', 0)} pass the frozen eligibility rules and {collection.get('invalid_total', 0)} do not.",
        f"Valid selected submissions: {gate['selected_valid_total']}/30. By form: {form_counts}.",
        f"Exclusions comprise assistance only={patterns.get('used_assistance', 0)}, technical issue only={patterns.get('technical_issue', 0)}, and both={patterns.get('used_assistance+technical_issue', 0)}.",
        f"The retained responses provide {coverage['labels']} item labels. No item has the required five valid labels across all forms, and complete five-label PairAcc is unavailable.",
        "",
        "## Data-Quality Diagnostics",
        "",
        f"Completion time ranges from {duration.get('minimum')} to {duration.get('maximum')} seconds (median {duration.get('median')} seconds); {duration.get('below_ten_minutes')} of {collection.get('mapped_total', 0)} submissions finish below ten minutes versus the planned 12--15 minutes.",
        f"Eligible label-level referent agreement with author gold is {exploratory['referent']['correct']}/{exploratory['referent']['total']} ({100 * exploratory['referent']['rate']:.1f}%).",
        f"Eligible label-level execution agreement is {exploratory['execution']['correct']}/{exploratory['execution']['total']} ({100 * exploratory['execution']['rate']:.1f}%).",
        "These unequal-form, incomplete-sample quantities are diagnostics only; they are not fixed-rater agreement or confirmatory endpoints.",
    ]
    relaxed = report.get("post_hoc_assistance_relaxed_sensitivity")
    if relaxed:
        relaxed_forms = " | ".join(
            f"{form}: {relaxed['selected_by_form'][form]}" for form in FORMS
        )
        lines.extend(
            [
                "",
                "## Post-Hoc Assistance-Relaxed Sensitivity",
                "",
                f"Ignoring the frozen assistance exclusion while retaining the technical-issue exclusion yields {relaxed['participants']} participants ({relaxed_forms}). Referent agreement is {relaxed['referent']['correct']}/{relaxed['referent']['total']} ({100 * relaxed['referent']['rate']:.1f}%) and execution agreement is {relaxed['execution']['correct']}/{relaxed['execution']['total']} ({100 * relaxed['execution']['rate']:.1f}%).",
                "The questionnaire combines translation, generative AI, search, and other-person help in one response, so this slice cannot isolate benign translation and remains non-evidentiary.",
            ]
        )
    primary = report.get("non_evidentiary_all_primary_sensitivity")
    if primary:
        lines.extend(
            [
                "",
                "## All-Primary Sensitivity",
                "",
                f"Using all 30 primary submissions despite frozen exclusions gives referent majority--gold {primary['referent']['majority_gold']}/72, execution majority--gold {primary['execution']['majority_gold']}/72, and changed-pair accuracy {primary['referent']['pair_majority_correct']}/18.",
                "This complete-looking matrix is reported only to show that relaxing eligibility does not rescue the construct audit; it is not evidence.",
            ]
        )
    lines.extend(
        [
            "",
            "## Interpretation and Follow-Up",
            "",
            "The collection cannot strengthen the TRI construct claim. It instead documents a failed eligibility and engagement gate.",
            "Additional eligibility-passing responses needed by form are "
            + ", ".join(f"{form}={needed[form]}" for form in FORMS)
            + ". Because the frozen allocation included only one reserve per form, any continuation beyond those slots requires a prospective protocol amendment before recruitment.",
            "A clean follow-up should retain the frozen items and gold, separate translation from AI/search/other-person assistance, add a comprehension gate before formal items, and recruit from a monitored panel. No new model experiment repairs this human-evidence gap.",
            "",
        ]
    )
    return "\n".join(lines)


def dump_report(report: dict[str, Any], path: Path) -> None:
    path.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
